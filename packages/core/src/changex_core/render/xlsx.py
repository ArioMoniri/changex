"""Non-native xlsx overlay renderer: a separate, annotated REVIEW copy.

Honesty (read this)
-------------------
xlsx has **no robust native track-changes** mechanism. There is therefore no
"accept/reject in place" the way Word owns docx revisions. The authoritative
record of *what changed, where, why, by whom* is the ``.changex`` journal; this
renderer projects that journal onto a **review surface**: a copy of the workbook
where directly-edited cells are colored, carry a threaded comment describing the
change, and a generated ``Changes`` audit sheet lists every op.

Non-destructive contract
-------------------------
:func:`build_review_workbook` takes the *clean* deliverable workbook bytes (the
exact recalculated result of applying the journal) plus the list of changes, and
returns the bytes of a **separate** review workbook. It never mutates the clean
deliverable and never injects the audit sheet into it — the clean file stays a
plain, shippable ``.xlsx``; the overlay is its annotated sibling.

Only **directly-edited** cells/formulas (the ops in the journal) are highlighted.
Cells whose values merely *recalculated* as dependents are not ops and are not
annotated — surfacing them would falsely attribute machine recomputation to the
agent.
"""

from __future__ import annotations

import io
from dataclasses import dataclass
from typing import TYPE_CHECKING

if TYPE_CHECKING:  # pragma: no cover - typing only
    from openpyxl.workbook.workbook import Workbook

# Fill / styling constants for the overlay. Kept here so the adapter stays focused
# on model/op mechanics and the renderer owns all presentation.
_VALUE_CHANGE_FILL = "FFFDE9A9"  # warm amber: a cell value was edited
_FORMULA_CHANGE_FILL = "FFCFE2F3"  # cool blue: a formula was edited
_ROW_INSERT_FILL = "FFD9EAD3"  # green: a row was inserted
_ROW_DELETE_FILL = "FFF4CCCC"  # red: a row was deleted (recorded on the audit sheet)
_AUDIT_SHEET_TITLE = "Changes"
_AUDIT_HEADER = ("Seq", "Kind", "Sheet", "Ref/At", "Before", "After", "Author")


@dataclass(frozen=True)
class CellChange:
    """One directly-edited cell to annotate on the review copy.

    ``kind`` is ``"cell.set"`` or ``"formula.set"``; ``ref`` is the A1 reference
    on ``sheet``. ``before`` / ``after`` are the stringified prior / new contents.
    """

    seq: int
    kind: str
    sheet: str
    ref: str
    before: str
    after: str
    author: str


@dataclass(frozen=True)
class RowChange:
    """One row-level structural change to record on the audit sheet.

    ``kind`` is ``"row.insert"`` or ``"row.delete"``; ``at`` is the 1-based row.
    For deletes, ``value`` is the comma-joined prior row contents (for replay
    visibility); inserts carry an empty ``value``.
    """

    seq: int
    kind: str
    sheet: str
    at: int
    value: str
    author: str


def _comment_text(change: CellChange) -> str:
    """Human-readable comment body for an edited cell."""
    verb = "Formula changed" if change.kind == "formula.set" else "Value changed"
    before = change.before if change.before != "" else "(empty)"
    after = change.after if change.after != "" else "(empty)"
    return f"ChangeX [{change.author}] seq {change.seq}: {verb}\n{before} -> {after}"


def build_review_workbook(
    clean_bytes: bytes,
    cell_changes: list[CellChange],
    row_changes: list[RowChange],
    *,
    inserted_rows: dict[str, set[int]] | None = None,
) -> bytes:
    """Return the bytes of an annotated REVIEW copy of ``clean_bytes``.

    The clean deliverable is loaded fresh into openpyxl (so the caller's clean
    bytes are never mutated), then directly-edited cells are filled + commented,
    inserted rows are tinted, and a ``Changes`` audit sheet is appended. The
    annotated workbook is serialized and returned as new bytes.

    Args:
        clean_bytes: The clean (recalculated) deliverable workbook bytes.
        cell_changes: Directly-edited cells (``cell.set`` / ``formula.set`` ops).
        row_changes: Row inserts/deletes for the audit sheet.
        inserted_rows: Optional ``{sheet: {row_index, ...}}`` of currently-present
            inserted rows to tint green in the body.

    Returns:
        New ``.xlsx`` bytes for the review copy. Empty when there are no changes
        still returns a valid annotated copy (just the audit sheet header).
    """
    from openpyxl import load_workbook
    from openpyxl.comments import Comment
    from openpyxl.styles import Font, PatternFill

    workbook: "Workbook" = load_workbook(io.BytesIO(clean_bytes))
    inserted = inserted_rows or {}

    for change in cell_changes:
        if change.sheet not in workbook.sheetnames:
            continue
        ws = workbook[change.sheet]
        fill_color = (
            _FORMULA_CHANGE_FILL if change.kind == "formula.set" else _VALUE_CHANGE_FILL
        )
        try:
            cell = ws[change.ref]
        except (ValueError, KeyError):  # malformed ref -> skip annotation, journal still holds it
            continue
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        cell.comment = Comment(_comment_text(change), f"ChangeX/{change.author}")

    for sheet_name, rows in inserted.items():
        if sheet_name not in workbook.sheetnames:
            continue
        ws = workbook[sheet_name]
        tint = PatternFill(start_color=_ROW_INSERT_FILL, end_color=_ROW_INSERT_FILL, fill_type="solid")
        max_col = max(ws.max_column, 1)
        for row_idx in rows:
            for col in range(1, max_col + 1):
                ws.cell(row=row_idx, column=col).fill = tint

    audit = workbook.create_sheet(title=_AUDIT_SHEET_TITLE)
    bold = Font(bold=True)
    for col, header in enumerate(_AUDIT_HEADER, start=1):
        c = audit.cell(row=1, column=col, value=header)
        c.font = bold
    audit_row = 2
    combined: list[tuple[int, list[object]]] = []
    for ch in cell_changes:
        combined.append((ch.seq, [ch.seq, ch.kind, ch.sheet, ch.ref, ch.before, ch.after, ch.author]))
    for rc in row_changes:
        combined.append(
            (rc.seq, [rc.seq, rc.kind, rc.sheet, rc.at, rc.value, "", rc.author])
        )
    for _seq, values in sorted(combined, key=lambda item: item[0]):
        for col, value in enumerate(values, start=1):
            audit.cell(row=audit_row, column=col, value=value)
        audit_row += 1

    out = io.BytesIO()
    workbook.save(out)
    return out.getvalue()


__all__ = ["CellChange", "RowChange", "build_review_workbook"]
