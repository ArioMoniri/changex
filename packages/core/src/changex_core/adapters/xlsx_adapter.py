"""The xlsx adapter: load -> sheet/row/cell model -> apply v0.2 ops -> overlay.

Honesty (read this)
-------------------
xlsx has **no robust native track-changes** — there is no Word-style
accept/reject baked into the file. So unlike :class:`DocxAdapter` (which emits
real ``w:ins`` / ``w:del`` revisions), this adapter's review surface is
**non-native**: the authoritative record is the ``.changex`` journal, and
:meth:`render_tracked` projects it onto a *separate* annotated REVIEW workbook
(colored cells + threaded comments + a generated ``Changes`` audit sheet). The
clean deliverable workbook is never mutated and never carries the audit sheet.

Identity strategy
-----------------
Cells are addressed by **natural key** ``"<sheet>!<ref>"`` (e.g. ``"Q3!B7"``) —
the spreadsheet analogue of a docx ``node_id``. Each row carries a stable,
adapter-minted ``rowId`` so that a cell *follows its row* under ``row.insert`` /
``row.delete``: the row's identity (and therefore its cells' identities) is the
``rowId``, while the A1 ``ref`` is recomputed from the row's *current* position.
This keeps replay deterministic even as rows shift.

What is and is not an op
------------------------
Only **directly-edited** cells/formulas are ops (``cell.set`` / ``formula.set``)
and only those are highlighted in the overlay. A cell whose value merely
*recalculated* because a dependent formula changed is **not** an op — attributing
machine recomputation to the agent would be dishonest. openpyxl does not evaluate
formulas, so the model stores the literal cell contents (value or ``=formula``).
"""

from __future__ import annotations

import io
from dataclasses import dataclass, field
from typing import TYPE_CHECKING, Any, Optional

from changex_core.adapters.base import (
    BeforeMismatchError,
    DocumentAdapter,
    NodeNotFoundError,
    OversizedOpError,
)
from changex_core.journal.canonical import sha256_hex
from changex_core.model.nodes import Node, NodeKind
from changex_core.ops.vocabulary import (
    CellSet,
    FormulaSet,
    Op,
    RowDelete,
    RowInsert,
)
from changex_core.paths import safe_path
from changex_core.render.xlsx import CellChange, RowChange, build_review_workbook

if TYPE_CHECKING:  # pragma: no cover - typing only
    from openpyxl.workbook.workbook import Workbook

DEFAULT_AUTHOR = "ChangeX agent"
DEFAULT_DATE = "2026-06-02T00:00:00Z"

# A row.delete may not drop a sheet that is mostly the row being removed in a
# single op when the sheet is tiny; we cap structural ops the same way docx caps
# text ops, so an agent cannot wipe a whole small sheet in one unreviewable op.
MAX_ROW_DELETE_FRACTION = 0.9


def _split_ref(ref: str) -> tuple[str, int]:
    """Split an A1 ``ref`` (``"B7"``) into ``(column_letters, row_number)``.

    Raises:
        BeforeMismatchError: if ``ref`` is not a simple single-cell A1 reference.
    """
    from openpyxl.utils import coordinate_to_tuple
    from openpyxl.utils.exceptions import InvalidFileException

    try:
        row, col = coordinate_to_tuple(ref)
    except (ValueError, TypeError, InvalidFileException) as exc:
        raise BeforeMismatchError(f"invalid cell ref {ref!r}: {exc}") from exc
    from openpyxl.utils import get_column_letter

    return get_column_letter(col), int(row)


@dataclass
class _Cell:
    """One modeled cell: its current literal content plus the directly-edited flag."""

    col: str  # column letters, e.g. "B"
    value: str  # literal contents: a value string, or "=formula"
    edited_kind: Optional[str] = None  # None | "cell.set" | "formula.set"
    edit_before: str = ""
    edit_seq: int = 0


@dataclass
class _Row:
    """One modeled row: a stable rowId and its cells keyed by column letters."""

    row_id: int
    cells: dict[str, _Cell] = field(default_factory=dict)
    inserted: bool = False


@dataclass
class _Sheet:
    """Adapter-side state for one worksheet (ordered rows + the base contents)."""

    name: str
    rows: list[_Row] = field(default_factory=list)


class XlsxAdapter(DocumentAdapter):
    """Loads a .xlsx, applies v0.2 cell/row ops, renders a non-native overlay.

    The journal (not the file) is the source of truth; :meth:`render_tracked`
    builds a separate annotated review copy. See the module docstring for the
    fidelity statement.
    """

    def __init__(
        self,
        raw_bytes: bytes,
        *,
        author: str = DEFAULT_AUTHOR,
        date: str = DEFAULT_DATE,
    ) -> None:
        self._raw = raw_bytes
        self._author = author
        self._date = date
        self._baseline_sha = sha256_hex(raw_bytes)
        self._sheets: list[_Sheet] = []
        self._row_seq = 0
        self._cell_changes: list[CellChange] = []
        self._row_changes: list[RowChange] = []
        self._build_model(raw_bytes)

    # -- construction ---------------------------------------------------------

    @classmethod
    def load(
        cls, path: str, *, author: str = DEFAULT_AUTHOR, date: str = DEFAULT_DATE, **_: Any
    ) -> "XlsxAdapter":
        """Load a .xlsx from a sanitized path (extra kwargs are accepted+ignored)."""
        resolved = safe_path(path, must_exist=True, allow_suffixes=(".xlsx",))
        raw = resolved.read_bytes()
        return cls(raw, author=author, date=date)

    def _mint_row_id(self) -> int:
        self._row_seq += 1
        return self._row_seq

    @staticmethod
    def _literal(cell: Any) -> str:
        """Return a cell's literal content as a string ("" for empty)."""
        value = cell.value
        if value is None:
            return ""
        return str(value)

    def _build_model(self, raw: bytes) -> None:
        from openpyxl import load_workbook

        workbook: "Workbook" = load_workbook(io.BytesIO(raw))
        self._sheets = []
        for name in workbook.sheetnames:
            ws = workbook[name]
            sheet = _Sheet(name=name)
            max_row = ws.max_row or 0
            max_col = ws.max_column or 0
            for r in range(1, max_row + 1):
                row = _Row(row_id=self._mint_row_id())
                from openpyxl.utils import get_column_letter

                for c in range(1, max_col + 1):
                    cell = ws.cell(row=r, column=c)
                    literal = self._literal(cell)
                    if literal != "":
                        col_letter = get_column_letter(c)
                        row.cells[col_letter] = _Cell(col=col_letter, value=literal)
                sheet.rows.append(row)
            self._sheets.append(sheet)

    # -- DocumentAdapter contract --------------------------------------------

    def baseline_sha256(self) -> str:
        return self._baseline_sha

    def _sheet(self, name: str) -> _Sheet:
        for sheet in self._sheets:
            if sheet.name == name:
                return sheet
        raise NodeNotFoundError(f"no sheet named {name!r}")

    def to_model(self) -> Node:
        """Return the model tree: root -> one PARAGRAPH node per non-empty cell.

        Node ids are natural keys ``"<sheet>!<ref>"``; the A1 ``ref`` is derived
        from each cell's *current* row position so addresses track row shifts.
        """
        root = Node(node_id="root", node_kind=NodeKind.DOCUMENT, path="/workbook")
        for sheet in self._sheets:
            for r_index, row in enumerate(sheet.rows, start=1):
                for col, cell in sorted(row.cells.items()):
                    ref = f"{col}{r_index}"
                    node_id = f"{sheet.name}!{ref}"
                    root.children.append(
                        Node(
                            node_id=node_id,
                            node_kind=NodeKind.PARAGRAPH,
                            path=f"/workbook/{sheet.name}/{ref}",
                            value=cell.value,
                            attrs={
                                "sheet": sheet.name,
                                "ref": ref,
                                "row_id": row.row_id,
                                "is_formula": cell.value.startswith("="),
                            },
                        )
                    )
        return root

    def set_model(self, root: Node) -> None:
        """Reset adapter state to ``root`` (used by :meth:`Journal.replay`)."""
        self._sheets = []
        self._row_seq = 0
        self._cell_changes = []
        self._row_changes = []
        # Group cell nodes by sheet then by row number, recreating rows in order.
        by_sheet: dict[str, dict[int, dict[str, str]]] = {}
        order: list[str] = []
        for node in root.children:
            if node.node_kind != NodeKind.PARAGRAPH:
                continue
            sheet_name = str(node.attrs.get("sheet", ""))
            ref = str(node.attrs.get("ref", ""))
            if not sheet_name or not ref:
                continue
            col, row_num = _split_ref(ref)
            if sheet_name not in by_sheet:
                by_sheet[sheet_name] = {}
                order.append(sheet_name)
            by_sheet[sheet_name].setdefault(row_num, {})[col] = str(node.value or "")
        for sheet_name in order:
            sheet = _Sheet(name=sheet_name)
            rows_map = by_sheet[sheet_name]
            for row_num in range(1, (max(rows_map) if rows_map else 0) + 1):
                row = _Row(row_id=self._mint_row_id())
                for col, literal in rows_map.get(row_num, {}).items():
                    if literal != "":
                        row.cells[col] = _Cell(col=col, value=literal)
                sheet.rows.append(row)
            self._sheets.append(sheet)

    def resolve(self, node_id: str) -> Node | None:
        return self.to_model().find(node_id)

    # -- apply ----------------------------------------------------------------

    def apply(self, op: Op) -> None:
        """Apply one v0.2 spreadsheet op, validating ``before`` and op size."""
        if isinstance(op, CellSet):
            self._apply_cell_set(op, kind="cell.set")
        elif isinstance(op, FormulaSet):
            self._apply_cell_set(op, kind="formula.set")
        elif isinstance(op, RowInsert):
            self._apply_row_insert(op)
        elif isinstance(op, RowDelete):
            self._apply_row_delete(op)
        else:  # pragma: no cover - exhaustive over xlsx ops
            raise TypeError(f"unsupported op type {type(op).__name__} for xlsx")

    def _locate_row(self, sheet: _Sheet, row_num: int) -> _Row:
        """Return the row at 1-based ``row_num``, extending the sheet if needed."""
        while len(sheet.rows) < row_num:
            sheet.rows.append(_Row(row_id=self._mint_row_id()))
        return sheet.rows[row_num - 1]

    def _next_seq(self) -> int:
        return len(self._cell_changes) + len(self._row_changes) + 1

    def _apply_cell_set(self, op: "CellSet | FormulaSet", *, kind: str) -> None:
        sheet = self._sheet(op.sheet)
        col, row_num = _split_ref(op.ref)
        row = self._locate_row(sheet, row_num)
        existing = row.cells.get(col)
        current = existing.value if existing is not None else ""
        if op.before != current:
            raise BeforeMismatchError(
                f"{kind} before {op.before!r} != current {current!r} at "
                f"{op.sheet}!{op.ref}"
            )
        if existing is None:
            existing = _Cell(col=col, value="")
            row.cells[col] = existing
        existing.value = op.after
        existing.edited_kind = kind
        existing.edit_before = op.before
        seq = self._next_seq()
        existing.edit_seq = seq
        self._cell_changes.append(
            CellChange(
                seq=seq,
                kind=kind,
                sheet=op.sheet,
                ref=f"{col}{row_num}",
                before=op.before,
                after=op.after,
                author=self._author,
            )
        )

    def _apply_row_insert(self, op: RowInsert) -> None:
        sheet = self._sheet(op.sheet)
        at = op.at
        if at < 1:
            raise BeforeMismatchError(f"row.insert at must be >= 1 (got {at})")
        pos = min(at - 1, len(sheet.rows))
        sheet.rows.insert(pos, _Row(row_id=self._mint_row_id(), inserted=True))
        self._row_changes.append(
            RowChange(
                seq=self._next_seq(),
                kind="row.insert",
                sheet=op.sheet,
                at=at,
                value="",
                author=self._author,
            )
        )

    def _apply_row_delete(self, op: RowDelete) -> None:
        sheet = self._sheet(op.sheet)
        at = op.at
        if at < 1 or at > len(sheet.rows):
            raise NodeNotFoundError(f"row.delete at {at} out of range on {op.sheet!r}")
        if len(sheet.rows) and len([r for r in sheet.rows if r.cells]) <= 1:
            # Deleting the only populated row of a sheet is an oversized structural op.
            if any(sheet.rows[at - 1].cells):
                raise OversizedOpError(
                    "split_required: this row.delete removes the sheet's only "
                    "populated row; review it as a sheet-level change instead."
                )
        removed = sheet.rows.pop(at - 1)
        captured = [removed.cells[c].value for c in sorted(removed.cells)]
        self._row_changes.append(
            RowChange(
                seq=self._next_seq(),
                kind="row.delete",
                sheet=op.sheet,
                at=at,
                value=", ".join(captured) or ", ".join(str(v) for v in op.value),
                author=self._author,
            )
        )

    # -- render / save --------------------------------------------------------

    def _clean_bytes(self) -> bytes:
        """Build the CLEAN (unannotated) deliverable workbook from the model.

        This is the shippable result of applying every active op: a plain .xlsx
        with the new cell contents and no overlay. Built fresh from openpyxl so
        the original bytes are never mutated.
        """
        from openpyxl import Workbook

        workbook = Workbook()
        # Workbook() starts with one default sheet; remove it so our sheets drive.
        default = workbook.active
        workbook.remove(default)
        for sheet in self._sheets:
            ws = workbook.create_sheet(title=sheet.name)
            for r_index, row in enumerate(sheet.rows, start=1):
                for col, cell in row.cells.items():
                    ws[f"{col}{r_index}"] = cell.value
        if not workbook.sheetnames:  # degenerate empty workbook
            workbook.create_sheet(title="Sheet1")
        return self._serialize(workbook)

    @staticmethod
    def _serialize(workbook: "Workbook") -> bytes:
        out = io.BytesIO()
        workbook.save(out)
        return out.getvalue()

    def _inserted_rows(self) -> dict[str, set[int]]:
        """Return ``{sheet: {1-based row index, ...}}`` for inserted rows."""
        result: dict[str, set[int]] = {}
        for sheet in self._sheets:
            rows = {i for i, row in enumerate(sheet.rows, start=1) if row.inserted}
            if rows:
                result[sheet.name] = rows
        return result

    def render_tracked(self) -> bytes:
        """Return the annotated NON-NATIVE REVIEW copy bytes (never the clean file).

        The clean deliverable is rebuilt from the model, then a *separate* copy is
        annotated (colored edited cells + comments + a ``Changes`` audit sheet).
        The clean deliverable itself is never mutated and never carries the audit
        sheet — it stays a plain, shippable workbook; this overlay is its sibling.
        """
        clean = self._clean_bytes()
        return build_review_workbook(
            clean,
            list(self._cell_changes),
            list(self._row_changes),
            inserted_rows=self._inserted_rows(),
        )

    def clean_workbook_bytes(self) -> bytes:
        """Return the CLEAN deliverable bytes (no overlay, no audit sheet).

        Exposed so callers that want the shippable workbook (not the review copy)
        can get it without the annotations :meth:`render_tracked` adds.
        """
        return self._clean_bytes()

    def save(self, out_path: str) -> None:
        """Save the annotated review workbook to a sanitized ``out_path``."""
        resolved = safe_path(out_path, allow_suffixes=(".xlsx",))
        resolved.parent.mkdir(parents=True, exist_ok=True)
        resolved.write_bytes(self.render_tracked())

    # -- accessors ------------------------------------------------------------

    def node_id_map(self) -> dict[str, str]:
        """Spreadsheets address by natural key, so there is no carrier map."""
        return {}
