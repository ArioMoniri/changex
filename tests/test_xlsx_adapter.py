"""XlsxAdapter op tests against a real generated ``.xlsx`` (M3).

These cover, via the public ``changex_core`` API + the registry:

* natural-key cell addressing (``"<sheet>!<ref>"``) and a stored ``row_id`` so a
  cell follows its row under ``row.insert`` (the M3 identity requirement);
* the four v0.2 spreadsheet ops apply (``cell.set`` / ``formula.set`` /
  ``row.insert`` / ``row.delete``) and the model projects the new contents;
* ONLY directly-edited cells are recorded as changes (a recalculated dependent is
  NOT an op) — encoded as: formula edits annotate, untouched cells do not;
* ``before``-mismatch and out-of-range guards refuse bad ops at the boundary;
* ``render_tracked`` builds a NON-DESTRUCTIVE review copy (colored cells +
  threaded comments + a generated ``Changes`` audit sheet) and the clean
  deliverable NEVER carries the audit sheet or annotations;
* the journal of those ops verifies and replays back to the live model.

xlsx has NO native track-changes: the journal + overlay are the review surface.
"""

from __future__ import annotations

import io
from pathlib import Path

import pytest

import changex_core as cx
from changex_core.adapters import load_adapter

openpyxl = pytest.importorskip("openpyxl", reason="openpyxl is required for the xlsx adapter")

from openpyxl import Workbook, load_workbook  # noqa: E402  (after importorskip)

from changex_core.adapters.xlsx_adapter import XlsxAdapter  # noqa: E402


# --------------------------------------------------------------------------- #
# fixtures / helpers
# --------------------------------------------------------------------------- #
@pytest.fixture()
def sample_xlsx(tmp_path: Path) -> Path:
    """A small two-column workbook on a sheet named ``Q3`` with a formula cell."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Q3"
    ws["A1"] = "Item"
    ws["B1"] = "Amount"
    ws["A2"] = "Widget"
    ws["B2"] = 100
    ws["A3"] = "Gadget"
    ws["B3"] = 200
    ws["C2"] = "=B2*1.1"
    out = tmp_path / "book.xlsx"
    wb.save(str(out))
    return out


def _prov(session_id: str) -> cx.Provenance:
    return cx.Provenance(
        ts="2026-06-02T00:00:00Z",
        session_id=session_id,
        agent="claude-opus-4-8",
        vendor="anthropic",
        provenance_source="declared",
    )


# --------------------------------------------------------------------------- #
# Addressing: natural key + row-following rowId
# --------------------------------------------------------------------------- #
def test_cells_addressed_by_natural_key(sample_xlsx: Path) -> None:
    adapter = load_adapter(str(sample_xlsx))
    assert isinstance(adapter, XlsxAdapter)
    model = adapter.to_model()
    b2 = model.find("Q3!B2")
    assert b2 is not None and b2.value == "100"
    assert b2.attrs["sheet"] == "Q3" and b2.attrs["ref"] == "B2"
    c2 = model.find("Q3!C2")
    assert c2 is not None and c2.attrs["is_formula"] is True


def test_cell_follows_its_row_under_row_insert(sample_xlsx: Path) -> None:
    """GATING (M3 identity): a row insert above a cell shifts its A1 ref, and the
    stored ``row_id`` stays bound to the same logical row."""
    adapter = load_adapter(str(sample_xlsx))
    gadget_before = adapter.to_model().find("Q3!A3")
    assert gadget_before is not None and gadget_before.value == "Gadget"
    original_row_id = gadget_before.attrs["row_id"]

    adapter.apply(cx.RowInsert(sheet="Q3", at=3))  # insert above Gadget

    after = adapter.to_model()
    # Gadget has moved from A3 to A4; its row_id is preserved (followed the row).
    assert after.find("Q3!A3") is None
    moved = after.find("Q3!A4")
    assert moved is not None and moved.value == "Gadget"
    assert moved.attrs["row_id"] == original_row_id


# --------------------------------------------------------------------------- #
# v0.2 op application
# --------------------------------------------------------------------------- #
def test_cell_set_updates_value_and_validates_before(sample_xlsx: Path) -> None:
    adapter = load_adapter(str(sample_xlsx))
    adapter.apply(cx.CellSet(sheet="Q3", ref="B2", before="100", after="125"))
    assert adapter.to_model().find("Q3!B2").value == "125"
    # A wrong `before` is refused at the boundary.
    with pytest.raises(cx.BeforeMismatchError):
        adapter.apply(cx.CellSet(sheet="Q3", ref="B2", before="100", after="999"))


def test_formula_set_updates_formula(sample_xlsx: Path) -> None:
    adapter = load_adapter(str(sample_xlsx))
    adapter.apply(cx.FormulaSet(sheet="Q3", ref="C2", before="=B2*1.1", after="=B2*1.2"))
    c2 = adapter.to_model().find("Q3!C2")
    assert c2.value == "=B2*1.2" and c2.attrs["is_formula"] is True


def test_row_insert_and_row_delete(sample_xlsx: Path) -> None:
    adapter = load_adapter(str(sample_xlsx))
    adapter.apply(cx.RowInsert(sheet="Q3", at=2))  # blank row 2; everything shifts down
    shifted = adapter.to_model().find("Q3!A3")
    assert shifted is not None and shifted.value == "Widget"

    # Delete that blank inserted row; Widget returns to row 2.
    adapter.apply(cx.RowDelete(sheet="Q3", at=2, value=[]))
    assert adapter.to_model().find("Q3!A2").value == "Widget"


def test_row_delete_out_of_range_is_refused(sample_xlsx: Path) -> None:
    adapter = load_adapter(str(sample_xlsx))
    with pytest.raises(cx.NodeNotFoundError):
        adapter.apply(cx.RowDelete(sheet="Q3", at=999, value=[]))


def test_unknown_sheet_is_refused(sample_xlsx: Path) -> None:
    adapter = load_adapter(str(sample_xlsx))
    with pytest.raises(cx.NodeNotFoundError):
        adapter.apply(cx.CellSet(sheet="Nope", ref="A1", before="", after="x"))


# --------------------------------------------------------------------------- #
# Non-destructive overlay: review copy is annotated; clean deliverable is not
# --------------------------------------------------------------------------- #
def test_render_tracked_review_copy_has_audit_sheet_and_annotations(sample_xlsx: Path) -> None:
    adapter = load_adapter(str(sample_xlsx), author="claude-opus-4-8")
    adapter.apply(cx.CellSet(sheet="Q3", ref="B2", before="100", after="125"))
    adapter.apply(cx.FormulaSet(sheet="Q3", ref="C2", before="=B2*1.1", after="=B2*1.2"))

    review = load_workbook(io.BytesIO(adapter.render_tracked()))
    assert "Changes" in review.sheetnames  # generated audit sheet present
    ws = review["Q3"]
    # The directly-edited value cell is colored and carries a threaded comment.
    assert ws["B2"].comment is not None and "125" in ws["B2"].comment.text
    assert ws["B2"].fill.fill_type == "solid"
    # The audit sheet lists both edits.
    audit_text = " ".join(
        str(c.value) for row in review["Changes"].iter_rows() for c in row if c.value
    )
    assert "cell.set" in audit_text and "formula.set" in audit_text


def test_clean_deliverable_never_carries_audit_sheet_or_annotations(sample_xlsx: Path) -> None:
    """The clean workbook must stay a plain, shippable file — no overlay leak."""
    adapter = load_adapter(str(sample_xlsx))
    adapter.apply(cx.CellSet(sheet="Q3", ref="B2", before="100", after="125"))

    clean = load_workbook(io.BytesIO(adapter.clean_workbook_bytes()))
    assert "Changes" not in clean.sheetnames  # audit sheet NOT injected
    cell = clean["Q3"]["B2"]
    assert cell.value == "125"  # edit applied to the clean file
    assert cell.comment is None  # but no annotation
    assert cell.fill.fill_type in (None, "none")  # and no fill


def test_only_directly_edited_cells_are_annotated_not_dependents(sample_xlsx: Path) -> None:
    """A formula edit annotates ONLY the formula cell; the referenced value cell
    (a 'dependent' in spirit) is untouched and stays un-annotated."""
    adapter = load_adapter(str(sample_xlsx))
    adapter.apply(cx.FormulaSet(sheet="Q3", ref="C2", before="=B2*1.1", after="=B2*1.2"))
    review = load_workbook(io.BytesIO(adapter.render_tracked()))
    ws = review["Q3"]
    assert ws["C2"].comment is not None  # the edited formula cell is annotated
    assert ws["B2"].comment is None  # the un-edited value cell is NOT


def test_save_writes_review_copy(sample_xlsx: Path, tmp_path: Path) -> None:
    adapter = load_adapter(str(sample_xlsx))
    adapter.apply(cx.CellSet(sheet="Q3", ref="B2", before="100", after="125"))
    out = tmp_path / "review.xlsx"
    adapter.save(str(out))
    assert "Changes" in load_workbook(str(out)).sheetnames


# --------------------------------------------------------------------------- #
# Journal integration: verify + replay back to the live model
# --------------------------------------------------------------------------- #
def test_journal_over_xlsx_verifies_and_replays(sample_xlsx: Path, tmp_path: Path) -> None:
    adapter = load_adapter(str(sample_xlsx), author="m")
    baseline = cx.Node.from_dict(adapter.to_model().to_dict())

    header = cx.Header.create(
        baseline_sha256=adapter.baseline_sha256(),
        filename=sample_xlsx.name,
        doc_format="xlsx",
    )
    journal = cx.Journal.open(str(tmp_path / "s.changex"), header=header)
    sid = journal.header.session_id

    def record(op: cx.Op) -> cx.Event:
        adapter.apply(op)
        node_id = cx.target_node_id(op) or ""
        node = adapter.resolve(node_id)
        target = cx.Target(
            node_id=node_id,
            node_kind=node.node_kind.value if node else "paragraph",
            path=node.path if node else "",
        )
        return journal.append(op, target, _prov(sid))

    record(cx.CellSet(sheet="Q3", ref="B2", before="100", after="125"))
    record(cx.RowInsert(sheet="Q3", at=3))
    record(cx.FormulaSet(sheet="Q3", ref="C2", before="=B2*1.1", after="=B2*1.2"))

    assert journal.verify().ok
    fresh = load_adapter(str(sample_xlsx), author="m")
    replayed = journal.replay(fresh, baseline)
    live = {n.node_id: n.value for n in adapter.to_model().children}
    assert {n.node_id: n.value for n in replayed.children} == live
